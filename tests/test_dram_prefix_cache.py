from __future__ import annotations

import hashlib
import http.server
import json
import os
import subprocess
import sys
import tempfile
import threading
import time
import unittest
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from xml.etree import ElementTree


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT = REPO_ROOT / "PrefixCache-HBM-DRAM-SSD" / "C5"
RUN_ALL = REPO_ROOT / "PrefixCache-HBM-DRAM-SSD" / "Run-all"
RUN_ALL_PAYLOAD = REPO_ROOT / "PrefixCache-HBM-DRAM-SSD" / "Run-all.payload"
RUN_ALL_BUILDER = (
    REPO_ROOT / "PrefixCache-HBM-DRAM-SSD" / "tools" / "build_run_all_wrapper.py"
)
FAKE_VLLM = REPO_ROOT / "tests" / "fake_vllm.py"

FAKE_RUNTIME_MANAGER = r'''
import os
import signal
import sys
import time
import urllib.request
from pathlib import Path

ssd_path = Path(sys.argv[1])
restart_url = sys.argv[2]
pid = os.getpid()
request_path = Path(f"/tmp/mooncake-benchmark-{pid}.request")
status_path = Path(f"/tmp/mooncake-benchmark-{pid}.status")
start_ticks = sys.argv[3]
generation = 1
restart_requested = False

def write_status(state, request_id):
    temporary = status_path.with_suffix(".status.tmp")
    temporary.write_text(
        f"state={state}\n"
        f"generation={generation}\n"
        f"request_id={request_id}\n"
        f"startup_pid={pid}\n"
        f"startup_start_ticks={start_ticks}\n"
        "master_pid=\n"
        "vllm_pid=\n"
        f"ssd_session_path={ssd_path}\n"
        "master_log=fake\n"
        "vllm_log=fake\n"
        "message=fake runtime\n",
        encoding="utf-8",
    )
    os.chmod(temporary, 0o600)
    temporary.replace(status_path)

def request_restart(_signum, _frame):
    global restart_requested
    restart_requested = True

request_path.write_text("", encoding="ascii")
os.chmod(request_path, 0o600)
signal.signal(signal.SIGUSR1, request_restart)
write_status("ready", "initial")
while True:
    signal.pause()
    if not restart_requested:
        continue
    restart_requested = False
    values = {}
    for line in request_path.read_text(encoding="ascii").splitlines():
        if "=" in line:
            key, value = line.split("=", 1)
            values[key] = value
    request_id = values.get("request_id", "")
    if values.get("command") != "restart" or values.get("expected_generation") != str(generation):
        write_status("failed", request_id)
        continue
    write_status("restarting", request_id)
    urllib.request.urlopen(urllib.request.Request(restart_url, data=b"", method="POST"), timeout=5).read()
    generation += 1
    write_status("ready", request_id)
'''

try:
    import openpyxl
except ImportError:  # The production scripts intentionally need only stdlib.
    openpyxl = None


class FakeState:
    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.local_queries = 0
        self.local_hits = 0
        self.external_queries = 0
        self.external_hits = 0
        self.keys = 0
        self.dram_hits = 0
        self.ssd_hits = 0
        self.dram_hit_bytes = 0
        self.ssd_hit_bytes = 0
        self.dram_allocated_bytes = 0
        self.ssd_allocated_bytes = 0
        self.r1_write_count = 0
        self.valid_gets = 0
        self.r4_tier = "ssd"
        self.block_stat_path: Path | None = None
        self.runtime_read_path: Path | None = None
        self.ssd_read_sectors = 0
        self.runtime_read_bytes = 0


def start_fake_runtime_manager(ssd_path: Path, restart_url: str):
    start_ticks = "fake-start-ticks"
    manager = subprocess.Popen(
        [sys.executable, "-c", FAKE_RUNTIME_MANAGER, str(ssd_path), restart_url, start_ticks],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )
    request_path = Path(f"/tmp/mooncake-benchmark-{manager.pid}.request")
    status_path = Path(f"/tmp/mooncake-benchmark-{manager.pid}.status")
    deadline = time.monotonic() + 5
    while time.monotonic() < deadline:
        if status_path.is_file() and "state=ready" in status_path.read_text(encoding="utf-8"):
            return manager, request_path, status_path, start_ticks
        if manager.poll() is not None:
            stdout, stderr = manager.communicate()
            raise RuntimeError(f"fake runtime manager exited: stdout={stdout!r} stderr={stderr!r}")
        time.sleep(0.05)
    manager.terminate()
    manager.communicate(timeout=5)
    raise RuntimeError("fake runtime manager did not become ready")


def stop_fake_runtime_manager(manager, request_path: Path, status_path: Path) -> None:
    if manager is not None and manager.poll() is None:
        manager.terminate()
        manager.communicate(timeout=5)
    request_path.unlink(missing_ok=True)
    status_path.unlink(missing_ok=True)


class DramHandler(http.server.BaseHTTPRequestHandler):
    state: FakeState

    def log_message(self, fmt, *args):
        pass

    def send_bytes(self, body: bytes, content_type="application/json") -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        path = self.path.split("?", 1)[0]
        if path == "/v1/models":
            body = json.dumps(
                {"data": [{"id": "fake-model", "root": "fake-model"}]}
            ).encode()
            self.send_bytes(body)
            return
        if path == "/metrics":
            with self.state.lock:
                text = (
                    f"vllm:prefix_cache_queries_total {self.state.local_queries}\n"
                    f"vllm:prefix_cache_hits_total {self.state.local_hits}\n"
                    "vllm:external_prefix_cache_queries_total "
                    f"{self.state.external_queries}\n"
                    "vllm:external_prefix_cache_hits_total "
                    f"{self.state.external_hits}\n"
                )
            self.send_bytes(text.encode(), "text/plain")
            return
        if path == "/mooncake/metrics":
            with self.state.lock:
                text = (
                    f"master_allocated_bytes {self.state.dram_allocated_bytes}\n"
                    f"master_allocated_file_size_bytes {self.state.ssd_allocated_bytes}\n"
                )
            self.send_bytes(text.encode(), "text/plain")
            return
        if path == "/metrics/summary":
            with self.state.lock:
                text = f"Mem Storage: fake | Keys: {self.state.keys}"
            self.send_bytes(text.encode(), "text/plain")
            return
        self.send_error(404)

    def do_POST(self):
        path = self.path.split("?", 1)[0]
        length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(length) if length else b""
        if path == "/reset_prefix_cache":
            self.send_bytes(b"true", "text/plain")
            return
        if path == "/test/restart":
            with self.state.lock:
                self.state.dram_allocated_bytes = 0
            self.send_bytes(b"true", "text/plain")
            return
        if path == "/tokenize":
            prompt = json.loads(body)["prompt"]
            words = prompt.split(" ")
            prefix = int(hashlib.sha256(words[0].encode()).hexdigest()[:8], 16)
            tokens = [prefix ^ index for index in range(len(words))]
            self.send_bytes(json.dumps({"tokens": tokens}).encode())
            return
        if path == "/v1/completions":
            request_id = self.headers.get("x-request-id", "")
            prompt_tokens = 16 * 1024
            cacheable_tokens = prompt_tokens - 128
            with self.state.lock:
                self.state.local_queries += prompt_tokens
                self.state.external_queries += prompt_tokens
                if "-r1-" in request_id:
                    self.state.keys += 1
                    self.state.dram_allocated_bytes += 8192
                    # Later P1s can share SSD-resident KV blocks from an
                    # earlier P1, so their new allocation can be smaller.
                    self.state.r1_write_count += 1
                    blocks = 126 if self.state.r1_write_count == 1 else 110
                    self.state.ssd_allocated_bytes += 8192 * blocks // 128
                elif "-r2-" in request_id:
                    self.state.local_hits += cacheable_tokens
                    self.state.external_hits += cacheable_tokens
                elif "-r3-" in request_id:
                    self.state.external_hits += cacheable_tokens
                    self.state.dram_hits += 1
                    self.state.dram_hit_bytes += cacheable_tokens
                    self.state.valid_gets += 1
                elif "-r4-" in request_id:
                    self.state.external_hits += cacheable_tokens
                    if self.state.r4_tier == "ssd":
                        self.state.ssd_hits += 1
                        self.state.ssd_hit_bytes += cacheable_tokens
                        self.state.ssd_read_sectors += 32
                        self.state.runtime_read_bytes += 16 * 1024
                        if self.state.block_stat_path is not None:
                            self.state.block_stat_path.write_text(
                                f"0 0 {self.state.ssd_read_sectors} 0 0 0 0 0 0 0 0\n",
                                encoding="ascii",
                            )
                        if self.state.runtime_read_path is not None:
                            self.state.runtime_read_path.write_text(
                                str(self.state.runtime_read_bytes), encoding="ascii"
                            )
                    else:
                        self.state.dram_hits += 1
                        self.state.dram_hit_bytes += cacheable_tokens
                        self.state.dram_allocated_bytes += 8192
                    self.state.valid_gets += 1
            self.send_bytes(b'{"choices":[{"text":"x"}]}')
            return
        self.send_error(404)


class DramPrefixCacheTest(unittest.TestCase):
    def test_run_all_wrapper_is_current_and_paste_sized(self) -> None:
        completed = subprocess.run(
            [sys.executable, str(RUN_ALL_BUILDER), "--check"],
            text=True,
            capture_output=True,
            check=False,
            cwd=REPO_ROOT,
        )
        self.assertEqual(completed.returncode, 0, msg=completed.stderr)
        run_all_text = RUN_ALL.read_text(encoding="utf-8")
        self.assertLess(len(run_all_text.encode("utf-8")), 32 * 1024)
        self.assertLessEqual(
            max(len(line.encode("utf-8")) for line in run_all_text.splitlines()),
            128,
        )
        self.assertGreater(len(RUN_ALL_PAYLOAD.read_bytes()), len(run_all_text.encode()))

    def test_c5_interleaves_stages_and_writes_one_workbook(self) -> None:
        state = FakeState()
        DramHandler.state = state
        server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), DramHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir)
                config = root / "mooncake.json"
                config.write_text(
                    json.dumps({"enable_ssd_offload": False}), encoding="utf-8"
                )
                env = os.environ.copy()
                env.update(
                    {
                        "VLLM_BIN": f"{sys.executable} {FAKE_VLLM}",
                        "FAKE_VLLM_SEND_REQUESTS": "1",
                        "FAKE_VLLM_OMIT_REQUEST_ID": "1",
                        "VB_PORT": str(server.server_address[1]),
                        "VB_MOONCAKE_METRICS_URL": (
                            f"http://127.0.0.1:{server.server_address[1]}"
                            "/metrics/summary"
                        ),
                        "MOONCAKE_CONFIG_PATH": str(config),
                        "VB_RUN_DIR": str(root / "results"),
                        "VB_BATCHES": "2",
                        "VB_EXPECT_CONCURRENCIES": "5",
                        "VB_STORE_STABLE_SECONDS": "0",
                        "VB_STORE_TIMEOUT_SECONDS": "5",
                    }
                )
                completed = subprocess.run(
                    ["bash", "-s"],
                    input=SCRIPT.read_text(encoding="utf-8"),
                    text=True,
                    capture_output=True,
                    check=False,
                    env=env,
                    cwd=REPO_ROOT,
                    timeout=60,
                )
                self.assertEqual(
                    completed.returncode,
                    0,
                    msg=f"stdout:\n{completed.stdout}\nstderr:\n{completed.stderr}",
                )

                result_dir = root / "results"
                manifest = json.loads(
                    (result_dir / "p1-manifest.json").read_text(encoding="utf-8")
                )
                requests = list(manifest["requests"].values())
                self.assertEqual(len(requests), 10)
                self.assertEqual(len({row["prompt_sha256"] for row in requests}), 10)
                self.assertEqual(
                    len({row["first_block_sha256"] for row in requests}), 10
                )

                payload = json.loads(
                    (result_dir / "C5-result.json").read_text(encoding="utf-8")
                )
                for stage in ("R1-warmup", "R2-hbm-cache", "R3-dram-cache"):
                    stage_summary = payload["stage_summary"][stage]
                    self.assertEqual(stage_summary["successful_requests"], 10)
                    self.assertEqual(stage_summary["success_rate_percent"], 100)
                    self.assertEqual(stage_summary["ttft_samples"], 10)
                r1_hashes = [row[9] for row in payload["stage_rows"]["R1-warmup"]]
                r2_hashes = [row[9] for row in payload["stage_rows"]["R2-hbm-cache"]]
                r3_hashes = [row[9] for row in payload["stage_rows"]["R3-dram-cache"]]
                self.assertEqual(r1_hashes, r2_hashes)
                self.assertEqual(r1_hashes, r3_hashes)

                workbook = result_dir / "vb-result.xlsx"
                with zipfile.ZipFile(workbook) as archive:
                    self.assertIsNone(archive.testzip())
                    names = archive.read("xl/workbook.xml").decode()
                    for sheet in (
                        "Summary",
                        "R1-warmup",
                        "R2-hbm-cache",
                        "R3-dram-cache",
                        "validation",
                        "P1-manifest",
                    ):
                        self.assertIn(f'name="{sheet}"', names)
                    self.assertIn("xl/sharedStrings.xml", archive.namelist())
                    for name in archive.namelist():
                        if name.endswith(".xml"):
                            ElementTree.fromstring(archive.read(name))
                        if name.startswith("xl/worksheets/"):
                            self.assertNotIn(b"inlineStr", archive.read(name))
                if openpyxl is not None:
                    loaded = openpyxl.load_workbook(workbook, read_only=True)
                    self.assertEqual(
                        loaded.sheetnames,
                        [
                            "Summary",
                            "R1-warmup",
                            "R2-hbm-cache",
                            "R3-dram-cache",
                            "validation",
                            "P1-manifest",
                        ],
                    )
                    summary_headers = [
                        cell.value
                        for cell in next(loaded["Summary"].iter_rows(max_row=1))
                    ]
                    self.assertEqual(
                        summary_headers,
                        [
                            "Concurrency",
                            "R1-Warmup-TTFT(s)",
                            "R2-HBM-TTFT(s)",
                            "R3-DRAM-TTFT(s)",
                            "R3/R2-Degradation(s)",
                            "R3/R2-Degradation(%)",
                            "Success/All",
                        ],
                    )
                    summary = loaded["Summary"]
                    self.assertEqual(summary.max_row, 2)
                    self.assertEqual(summary.cell(2, 1).value, "C5")
                    self.assertAlmostEqual(
                        summary.cell(2, 5).value,
                        summary.cell(2, 4).value - summary.cell(2, 3).value,
                    )
                    self.assertAlmostEqual(
                        summary.cell(2, 6).value,
                        (summary.cell(2, 4).value / summary.cell(2, 3).value - 1)
                        * 100,
                    )
                    self.assertEqual(summary.cell(2, 7).value, "30/30")
                    forbidden = ("sha", "identity", "validation", "cache")
                    self.assertFalse(
                        any(
                            marker in str(header).lower()
                            for header in summary_headers
                            for marker in forbidden
                        )
                    )
                    loaded.close()
        finally:
            server.shutdown()
            server.server_close()

    def test_run_all_validates_hbm_dram_and_ssd_in_one_session(self) -> None:
        state = FakeState()
        DramHandler.state = state
        server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), DramHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        manager_info = None
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir)
                config = root / "mooncake.json"
                ssd_path = root / "ssd"
                ssd_path.mkdir()
                (ssd_path / "bucket-0").write_bytes(b"fake-ssd-kv")
                manager_info = start_fake_runtime_manager(
                    ssd_path,
                    f"http://127.0.0.1:{server.server_address[1]}/test/restart",
                )
                manager, request_path, status_path, start_ticks = manager_info
                block_stat = root / "ssd-stat"
                block_stat.write_text("0 0 0 0 0 0 0 0 0 0 0\n", encoding="ascii")
                runtime_read = root / "runtime-read-bytes"
                runtime_read.write_text("0", encoding="ascii")
                state.block_stat_path = block_stat
                state.runtime_read_path = runtime_read
                config.write_text(
                    json.dumps(
                        {
                            "enable_ssd_offload": True,
                            "ssd_offload_path": str(ssd_path),
                            "benchmark_ssd_verified": True,
                            "benchmark_ssd_kname": "fake0",
                            "benchmark_ssd_direct_io": False,
                            "benchmark_ssd_page_cache_drop": True,
                            "benchmark_ssd_io_mode": "posix-fadvise-dontneed",
                            "benchmark_startup_pid": manager.pid,
                            "benchmark_startup_start_ticks": start_ticks,
                            "benchmark_runtime_control": "signal-request-v1",
                            "benchmark_runtime_request_path": str(request_path),
                            "benchmark_runtime_status_path": str(status_path),
                        }
                    ),
                    encoding="utf-8",
                )
                env = os.environ.copy()
                env.pop("VB_RUN_DIR", None)
                env.update(
                    {
                        "VLLM_BIN": f"{sys.executable} {FAKE_VLLM}",
                        "FAKE_VLLM_SEND_REQUESTS": "1",
                        "VB_PORT": str(server.server_address[1]),
                        "VB_MOONCAKE_METRICS_URL": (
                            f"http://127.0.0.1:{server.server_address[1]}"
                            "/metrics/summary"
                        ),
                        "VB_MOONCAKE_PROMETHEUS_URL": (
                            f"http://127.0.0.1:{server.server_address[1]}"
                            "/mooncake/metrics"
                        ),
                        "MOONCAKE_CONFIG_PATH": str(config),
                        "VB_SSD_BLOCK_STAT_PATH": str(block_stat),
                        "VB_RUNTIME_READ_BYTES_PATH": str(runtime_read),
                        "VB_RUNTIME_START_TICKS_OVERRIDE": start_ticks,
                        "VB_BATCHES": "1",
                        "VB_STORE_STABLE_SECONDS": "0",
                        "VB_STORE_TIMEOUT_SECONDS": "5",
                        "VB_RUNTIME_RESTART_TIMEOUT_SECONDS": "30",
                        "TZ": "UTC",
                    }
                )
                run_all_text = RUN_ALL.read_text(encoding="utf-8")
                self.assertLess(len(run_all_text.encode("utf-8")), 32 * 1024)
                self.assertLessEqual(
                    max(len(line.encode("utf-8")) for line in run_all_text.splitlines()),
                    1024,
                )
                cn_timezone = timezone(timedelta(hours=8))
                started_cn = datetime.now(cn_timezone) - timedelta(seconds=1)
                completed = subprocess.run(
                    ["bash", "-s"],
                    input=run_all_text,
                    text=True,
                    capture_output=True,
                    check=False,
                    env=env,
                    cwd=root,
                    timeout=120,
                )
                self.assertEqual(
                    completed.returncode,
                    0,
                    msg=f"stdout:\n{completed.stdout}\nstderr:\n{completed.stderr}",
                )
                self.assertIn("SSD副本 | 新增", completed.stdout)
                self.assertNotIn("not stable", completed.stderr)

                result_roots = list((root / "vb-result").iterdir())
                self.assertEqual(len(result_roots), 1)
                self.assertRegex(
                    result_roots[0].name,
                    r"^\d{4}-\d{2}-\d{2} \d{2}-\d{2}-\d{2}(?:-\d{2})?$",
                )
                result_time = datetime.strptime(
                    result_roots[0].name[:19], "%Y-%m-%d %H-%M-%S"
                ).replace(tzinfo=cn_timezone)
                self.assertGreaterEqual(result_time, started_cn)
                self.assertLessEqual(
                    result_time, datetime.now(cn_timezone) + timedelta(seconds=1)
                )
                result_dir = result_roots[0]
                self.assertEqual(
                    {path.name for path in result_dir.glob("C*-result.json")},
                    {"C1-result.json", "C5-result.json", "C10-result.json", "C100-result.json"},
                )
                session = json.loads(
                    (result_dir / "session.json").read_text(encoding="utf-8")
                )
                self.assertEqual(session["completed_concurrencies"], [1, 5, 10, 100])
                self.assertEqual(session["storage_tiers"], ["HBM", "DRAM", "SSD"])
                self.assertEqual(session["ssd_transition"]["mode"], "signal-request-v1")
                runtime_status = dict(
                    line.split("=", 1)
                    for line in status_path.read_text(encoding="utf-8").splitlines()
                    if "=" in line
                )
                self.assertEqual(runtime_status["state"], "ready")
                self.assertEqual(runtime_status["generation"], "5")

                c1_payload = json.loads(
                    (result_dir / "C1-result.json").read_text(encoding="utf-8")
                )
                sources = {
                    row[2]: (row[19], row[20])
                    for row in c1_payload["validation_rows"]
                }
                self.assertEqual(sources["R1-warmup"], ("MISS", "yes"))
                self.assertEqual(sources["R2-hbm-cache"], ("HBM", "yes"))
                self.assertEqual(sources["R3-dram-cache"], ("DRAM", "yes"))
                self.assertEqual(sources["R4-ssd-cache"], ("SSD", "yes"))

                manifest = json.loads(
                    (result_dir / "p1-manifest.json").read_text(encoding="utf-8")
                )
                requests = list(manifest["requests"].values())
                self.assertEqual(len(requests), 116)
                self.assertEqual(len({row["prompt_sha256"] for row in requests}), 116)
                self.assertEqual(
                    len({row["first_block_sha256"] for row in requests}), 116
                )

                workbook = result_dir / "vb-result.xlsx"
                with zipfile.ZipFile(workbook) as archive:
                    self.assertIsNone(archive.testzip())
                    names = archive.read("xl/workbook.xml").decode()
                    self.assertEqual(names.count("<sheet "), 7)
                    self.assertIn('name="R4-ssd-cache"', names)
                    self.assertIn("xl/sharedStrings.xml", archive.namelist())
                    for name in archive.namelist():
                        if name.endswith(".xml"):
                            ElementTree.fromstring(archive.read(name))
                        if name.startswith("xl/worksheets/"):
                            self.assertNotIn(b"inlineStr", archive.read(name))
                if openpyxl is not None:
                    loaded = openpyxl.load_workbook(workbook, read_only=True)
                    summary = loaded["Summary"]
                    self.assertEqual(summary.max_row, 5)
                    self.assertEqual(
                        [summary.cell(row=row, column=1).value for row in range(2, 6)],
                        ["C1", "C5", "C10", "C100"],
                    )
                    self.assertEqual(
                        [summary.cell(row=row, column=12).value for row in range(2, 6)],
                        ["4/4", "20/20", "40/40", "400/400"],
                    )
                    loaded.close()
        finally:
            if manager_info is not None:
                stop_fake_runtime_manager(*manager_info[:3])
            server.shutdown()
            server.server_close()

    def test_run_all_rejects_r4_when_mooncake_reports_dram(self) -> None:
        state = FakeState()
        state.r4_tier = "dram"
        DramHandler.state = state
        server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), DramHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        manager_info = None
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                root = Path(temp_dir)
                ssd_path = root / "ssd"
                ssd_path.mkdir()
                (ssd_path / "bucket-0").write_bytes(b"fake-ssd-kv")
                manager_info = start_fake_runtime_manager(
                    ssd_path,
                    f"http://127.0.0.1:{server.server_address[1]}/test/restart",
                )
                manager, request_path, status_path, start_ticks = manager_info
                block_stat = root / "ssd-stat"
                block_stat.write_text("0 0 0 0 0 0 0 0 0 0 0\n", encoding="ascii")
                runtime_read = root / "runtime-read-bytes"
                runtime_read.write_text("0", encoding="ascii")
                state.block_stat_path = block_stat
                state.runtime_read_path = runtime_read
                config = root / "mooncake.json"
                config.write_text(
                    json.dumps(
                        {
                            "enable_ssd_offload": True,
                            "ssd_offload_path": str(ssd_path),
                            "benchmark_ssd_verified": True,
                            "benchmark_ssd_kname": "fake0",
                            "benchmark_ssd_direct_io": False,
                            "benchmark_ssd_page_cache_drop": True,
                            "benchmark_ssd_io_mode": "posix-fadvise-dontneed",
                            "benchmark_startup_pid": manager.pid,
                            "benchmark_startup_start_ticks": start_ticks,
                            "benchmark_runtime_control": "signal-request-v1",
                            "benchmark_runtime_request_path": str(request_path),
                            "benchmark_runtime_status_path": str(status_path),
                        }
                    ),
                    encoding="utf-8",
                )
                env = os.environ.copy()
                env.update(
                    {
                        "VLLM_BIN": f"{sys.executable} {FAKE_VLLM}",
                        "FAKE_VLLM_SEND_REQUESTS": "1",
                        "VB_PORT": str(server.server_address[1]),
                        "VB_MOONCAKE_METRICS_URL": (
                            f"http://127.0.0.1:{server.server_address[1]}"
                            "/metrics/summary"
                        ),
                        "VB_MOONCAKE_PROMETHEUS_URL": (
                            f"http://127.0.0.1:{server.server_address[1]}"
                            "/mooncake/metrics"
                        ),
                        "MOONCAKE_CONFIG_PATH": str(config),
                        "VB_SSD_BLOCK_STAT_PATH": str(block_stat),
                        "VB_RUNTIME_READ_BYTES_PATH": str(runtime_read),
                        "VB_RUNTIME_START_TICKS_OVERRIDE": start_ticks,
                        "VB_RUN_DIR": str(root / "results"),
                        "VB_BATCHES": "1",
                        "VB_EXPECT_CONCURRENCIES": "1",
                        "VB_RUN_CONCURRENCIES": "1",
                        "VB_STORE_STABLE_SECONDS": "0",
                        "VB_STORE_TIMEOUT_SECONDS": "5",
                        "VB_RUNTIME_RESTART_TIMEOUT_SECONDS": "30",
                    }
                )
                run_all_text = RUN_ALL.read_text(encoding="utf-8")
                completed = subprocess.run(
                    ["bash", "-s"],
                    input=run_all_text,
                    text=True,
                    capture_output=True,
                    check=False,
                    env=env,
                    cwd=REPO_ROOT,
                    timeout=60,
                )
                self.assertEqual(completed.returncode, 2)
                self.assertIn("R4 DRAM", completed.stdout)
                payload = json.loads(
                    (root / "results" / "C1-result.json").read_text(encoding="utf-8")
                )
                self.assertIsNone(
                    payload["stage_summary"]["R4-ssd-cache"]["avg_ttft_seconds"]
                )
                r4_validation = next(
                    row
                    for row in payload["validation_rows"]
                    if row[2] == "R4-ssd-cache"
                )
                self.assertEqual(r4_validation[20], "no")
        finally:
            if manager_info is not None:
                stop_fake_runtime_manager(*manager_info[:3])
            server.shutdown()
            server.server_close()


if __name__ == "__main__":
    unittest.main()
